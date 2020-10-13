#!/usr/bin/env python
# -*- coding: utf-8 -*-
# release  21-05-2017
import sys
from optparse import OptionParser
import os
from uadb import UaDb
import re
from openpyxl import Workbook
from decimal import *
import decimal
import pdb

def do_main(ini, file_xls, table):
    # rgx = re.compile(r'^\-?\d+\.\d*$')
    rgx = re.compile(r'^\-?\d+\.?\d*$')
    uadb = UaDb.from_file(ini)
    sql = " select * from %s " % (table)
    rt = uadb.fetchall(sql, [])
    cols = rt.cols
    rows = rt.rowset

    wb = Workbook()
    ws = wb.active

    for c, col in enumerate(cols):
        ws.cell(row=1, column=c + 1).value = col

    for r, row in enumerate(rows):
        for c, col in enumerate(cols):
            value = row[c]
            if value is None:
                ws.cell(row=r + 2, column=c + 1).value = value
                continue
            s = str(value)
            if rgx.match(s) is None:
                try:
                    v = re.sub(r'[^\x00-\x7F]', ' ', value)
                    ws.cell(row=r + 2, column=c + 1).value = v
                except Exception as err:
                    print(err)
            else:
                try:
                    s = Decimal(value)
                except decimal.InvalidOperation:
                    print("numeric ERROR! row:%s  col:%s value:%s " % (r, col, value))
                    s = value
                ws.cell(row=r + 2, column=c + 1).value = s

    wb.save(file_xls)
    os.chmod(file_xls, 0o666)


# campi numerici es: numerics ="4,8,11"
def do_main_num(ini, file_xls, table, numerics):
    uadb = UaDb.from_file(ini)
    sql = " select * from %s " % (table)
    rt = uadb.fetchall(sql, [])
    cols = rt.cols
    rows = rt.rowset
    wb = Workbook()
    ws = wb.active
    nums = [int(x) - 1 for x in numerics.split(',')]

    for c, col in enumerate(cols):
        ws.cell(row=1, column=c + 1).value = col

    for r, row in enumerate(rows):
        for c, col in enumerate(cols):
            value = row[c]
            if c in nums:
                if value is None:
                    s = 0
                elif str(value).lower() == 'null':
                    s = 0
                elif str(value) == '':
                    s = 0
                else:
                    try:
                        s = Decimal(value)
                    except decimal.InvalidOperation:
                        print("numeric ERROR! row:%s  col:%s value:%s " % (r, col, value))
                        s = 0
                # ws.cell(row=r + 2, column=c + 1).number_format = '#,###.00'
            else:
                # s = value
                s = re.sub(r'[^\x00-\x7F]', ' ', value)
            # print("r: %s c:%s val:%s " % (r, c, s))
            ws.cell(row=r + 2, column=c + 1).value = s
    wb.save(file_xls)
    os.chmod(file_xls, 0o666)


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-i", "--ini")
    parser.add_option("-x", "--xls")
    parser.add_option("-t", "--table")
    parser.add_option("-n", "--num")
    try:
        opts, args = parser.parse_args()
    except Exception:
        opts = None
    if opts is None or opts.xls is None or opts.ini is None or opts.table is None:
        print ("-i <db.ini> -x <file.xls> -t <table> [ -n \"1,6,..\" ]")
        sys.exit(0)
    # print("%s %s %s %s" % (opts.ini, opts.xls, opts.table, opts.num))
    if opts.num is None:
        do_main(opts.ini, opts.xls, opts.table)
    else:
        do_main_num(opts.ini, opts.xls, opts.table, opts.num)
