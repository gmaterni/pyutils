#!/usr/bin/env python
# -*- coding: utf-8 -*-
# release 15-03-2017

import sys
from optparse import OptionParser
import csv
from openpyxl import Workbook
import os
import decimal
# from pdb import set_trace
import datetime


def read_csv(csv_path, delimiter, doublequote):
    csv_file = open(csv_path, 'r+')
    csv_reader = csv.reader(csv_file, delimiter=delimiter, doublequote=doublequote)
    rows = []
    for row in csv_reader:
        rows.append(row)
    return rows


def read_fmt(fmt_path, delimiter):
    if fmt_path is None or len(fmt_path) == 0:
        return []
    with open(fmt_path, 'r+') as f:
        rows = f.readlines()
    fmts = [x.strip() for x in rows[0].split(delimiter)]
    return fmts


# cols_type: "1:f,2:d,3:i,.."
def write_xlsx(rows, xls_path, fmts):
    wb = Workbook()
    ws = wb.active
    cols = rows[0]
    if fmts == []:
        fmts = ['s' for c in cols]
    for c, col in enumerate(cols):
        ws.cell(row=1, column=c + 1).value = col
    for r, row in enumerate(rows[1:]):
        for c, col in enumerate(cols):
            val = row[c]
            t = fmts[c]
            # print("r:%s  c:%s  v:%s t:%s" % (r, c, val, t))
            if val is None or val.lower() == 'null' or val.lower() == 'none' or val.strip() == "":
                val = ""
            s = val
            le = len(val)
            if t == 'f' and le > 0:  # float
                try:
                    v = val.replace(',', '.')
                    s = decimal.Decimal(v)
                except decimal.InvalidOperation:
                    print("decimal ERROR! row:%s   c:%s  col:%s  value:%s  type:%s" % (r, c, col, val, t))
                    s = ""
            elif t == 'i' and le > 0:  # integer
                try:
                    s = int(val)
                except Exception:
                    print("integer ERROR! row:%s  c:%s  col:%s  value:%s  type:%s" % (r, c, col, val, t))
                    s = ""
            elif t == 'dmy' and le == 10:  # date eu
                try:
                    s = val.replace('-', '/')
                    dmy = [int(x) for x in s.split('/')]
                    d = datetime.date(dmy[2], dmy[1], dmy[0])
                    # s = d.strftime("%d/%m/%Y")
                    s = d
                except Exception:
                    print("date ERROR! row:%s  c:%s  col:%s value:%s  type:%s" % (r, c, col, val, t))
                    s = ""
            ws.cell(row=r + 2, column=c + 1).value = s
    wb.save(xls_path)
    os.chmod(xls_path, 0o666)


def do_main(csv_path, delimiter, doublequote, xls_path, fmt_path):
    rows = read_csv(csv_path, delimiter, doublequote)
    fmts = read_fmt(fmt_path, delimiter)
    write_xlsx(rows, xls_path, fmts)


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-c", "--csv")
    parser.add_option("-s", "--sep")
    parser.add_option("-x", "--xls")
    parser.add_option("-t", "--type")
    try:
        opts, args = parser.parse_args()
    except Exception:
        opts = None
    if opts is None or opts.csv is None or opts.xls is None:
        print ("-c <file.csv> -s <separator> -x <file.xlsx>  [-t \"1:f,j:s,k:i,..n:dmy\" ] (f)loat/i)int/s)string/dmy)date) ")
        sys.exit(0)
    opts.sep = '|' if opts.sep is None else opts.sep
    doublequote = False
    do_main(opts.csv, opts.sep, doublequote, opts.xls, opts.type)
