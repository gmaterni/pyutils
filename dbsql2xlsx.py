#!/usr/bin/env python
# -*- coding: utf-8 -*-
# release  21-05-2017
import sys
from optparse import OptionParser
import os
from uadb import UaDb
import re
from openpyxl import Workbook
from decimal import *


def do_main(ini, file_xls, sql):
    rgx = re.compile(r'^\-?\d+\.?\d*$')
    uadb = UaDb.from_file(ini)
    rt = uadb.fetchall(sql, [])
    cols = rt.cols
    rows = rt.rowset
    wb = Workbook()
    ws = wb.active
    for c, col in enumerate(cols):
        ws.cell(row=1, column=c + 1).value = col
    for r, row in enumerate(rows):
        for c, col in enumerate(cols):
            value = row[c]
            if value is None:
                ws.cell(row=r + 2, column=c + 1).value = value
                continue
            s = str(value)
            if rgx.match(s) is None:
                ws.cell(row=r + 2, column=c + 1).value = value
            else:
                try:
                    s = Decimal(value)
                    ws.cell(row=r + 2, column=c + 1).value = s
                except Exception as e:
                    print(e)
                    print(value)
                    ws.cell(row=r + 2, column=c + 1).value = value
    wb.save(file_xls)
    os.chmod(file_xls, 0o666)


def do_main_num(ini, file_xls, sql, cols_numeric):
    uadb = UaDb.from_file(ini)
    rt = uadb.fetchall(sql, [])
    cols = rt.cols
    rows = rt.rowset
    wb = Workbook()
    ws = wb.active
    nums = [cols.index(c) for c in cols_numeric.split(',')]
    for c, col in enumerate(cols):
        ws.cell(row=1, column=c + 1).value = col
    for r, row in enumerate(rows):
        for c, col in enumerate(cols):
            value = row[c]
            if c in nums:
                if value is None:
                    s = 0
                elif str(value).lower() == 'null':
                    s = 0
                elif str(value) == '':
                    s = 0
                else:
                    try:
                        s = Decimal(value)
                    except InvalidOperation:
                        print("numeric ERROR! row:%s  col:%s value:%s " % (r, col, value))
                        s = 0
                # ws.cell(row=r + 2, column=c + 1).number_format = '#,###.00'
            else:
                # s = value
                s = re.sub(r'[^\x00-\x7F]', ' ', value)
            ws.cell(row=r + 2, column=c + 1).value = s
    wb.save(file_xls)
    os.chmod(file_xls, 0o666)


if __name__ == "__main__":
    parser = OptionParser()
    parser.add_option("-i", "--ini")
    parser.add_option("-x", "--xls")
    parser.add_option("-s", "--sql")
    parser.add_option("-f", "--filesql")
    parser.add_option("-n", "--num")
    try:
        opts, args = parser.parse_args()
    except Exception:
        opts = None
    if opts is None or opts.xls is None or opts.ini is None:
        print (
            "-i <db.ini> -x <file.xls> -s <\"sql\"> -f < file.sql > [-n \"col_i, , col_j, ..\" ")
        sys.exit(0)
    sql = opts.sql
    if opts.filesql is not None:
        with open(opts.filesql, "r+") as f:
            sql = f.read()
    if opts.num is None:
        do_main(opts.ini, opts.xls, sql)
    else:
        do_main_num(opts.ini, opts.xls, sql, opts.num)
